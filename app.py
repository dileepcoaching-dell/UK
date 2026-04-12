import streamlit as st
import pandas as pd
from pathlib import Path
import PyPDF2
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import tempfile
import os

st.set_page_config(
    page_title="GSTR-3B Consolidator",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS
st.markdown("""
<style>
    * {
        margin: 0;
        padding: 0;
        box-sizing: border-box;
    }
    
    body {
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        min-height: 100vh;
    }
    
    .main {
        background-color: #f8f9fa;
        padding: 20px;
    }
    
    .stApp {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    }
    
    .header-title {
        color: white;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.3);
        font-size: 2.5em;
        font-weight: 700;
        margin-bottom: 10px;
        animation: slideDown 0.8s ease-out;
    }
    
    .header-subtitle {
        color: rgba(255,255,255,0.9);
        font-size: 1.1em;
        margin-bottom: 30px;
    }
    
    @keyframes slideDown {
        from {
            opacity: 0;
            transform: translateY(-20px);
        }
        to {
            opacity: 1;
            transform: translateY(0);
        }
    }
    
    .upload-section {
        background: white;
        padding: 30px;
        border-radius: 12px;
        box-shadow: 0 10px 30px rgba(0,0,0,0.1);
        margin-bottom: 30px;
        border-left: 5px solid #667eea;
    }
    
    .success-message {
        background: #d4edda;
        color: #155724;
        padding: 15px;
        border-radius: 8px;
        margin: 10px 0;
        border-left: 4px solid #28a745;
    }
    
    .error-message {
        background: #f8d7da;
        color: #721c24;
        padding: 15px;
        border-radius: 8px;
        margin: 10px 0;
        border-left: 4px solid #f5c6cb;
    }
    
    .info-section {
        background: #e7f3ff;
        color: #004085;
        padding: 15px;
        border-radius: 8px;
        margin: 10px 0;
        border-left: 4px solid #0056b3;
    }
    
    .download-btn {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 12px 30px;
        border-radius: 8px;
        text-align: center;
        font-weight: 600;
        margin-top: 10px;
        cursor: pointer;
        transition: transform 0.3s ease, box-shadow 0.3s ease;
    }
    
    .download-btn:hover {
        transform: translateY(-2px);
        box-shadow: 0 10px 20px rgba(102, 126, 234, 0.3);
    }
    
    .stats-card {
        background: white;
        padding: 20px;
        border-radius: 12px;
        box-shadow: 0 5px 15px rgba(0,0,0,0.08);
        margin: 10px 0;
        border-top: 4px solid #667eea;
    }
    
    .stats-value {
        font-size: 1.8em;
        font-weight: 700;
        color: #667eea;
        margin: 10px 0;
    }
    
    .stats-label {
        color: #666;
        font-size: 0.9em;
    }
</style>
""", unsafe_allow_html=True)

def extract_gstr3b_data(pdf_file):
    """Extract data from GSTR-3B PDF"""
    try:
        pdf_reader = PyPDF2.PdfReader(pdf_file)
        text = ""
        for page in pdf_reader.pages:
            text += page.extract_text()
        
        # Extract month/period
        period_match = re.search(r'Period\s+([A-Za-z]+)', text)
        period = period_match.group(1) if period_match else "Unknown"
        
        # Extract GSTIN
        gstin_match = re.search(r'GSTIN of the supplier\s+([A-Z0-9]+)', text)
        gstin = gstin_match.group(1) if gstin_match else ""
        
        # Extract outward supplies data
        data = {
            'Period': period,
            'GSTIN': gstin,
        }
        
        # Extract numerical values using regex patterns
        outward_other = re.search(r'Outward taxable supplies.*?\n\s*([0-9,]+\.?\d*)', text)
        outward_zero = re.search(r'Outward taxable supplies \(zero rated\).*?\n\s*([0-9,]+\.?\d*)', text)
        outward_nil = re.search(r'Other outward supplies.*?\n\s*([0-9,]+\.?\d*)', text)
        inward_rc = re.search(r'Inward supplies \(liable to reverse charge\).*?\n\s*([0-9,]+\.?\d*)', text)
        
        # Extract tax amounts
        igst_match = re.search(r'Integrated tax\s+([0-9,]+\.?\d*)', text)
        cgst_match = re.search(r'Central tax\s+([0-9,]+\.?\d*)', text)
        sgst_match = re.search(r'State/UT tax\s+([0-9,]+\.?\d*)', text)
        
        # Clean and convert values
        def clean_number(match_obj):
            if match_obj:
                val = match_obj.group(1).replace(',', '')
                try:
                    return float(val)
                except:
                    return 0.0
            return 0.0
        
        data['Outward Taxable (Other)'] = clean_number(outward_other)
        data['Outward Taxable (Zero Rated)'] = clean_number(outward_zero)
        data['Other Outward (Nil/Exempt)'] = clean_number(outward_nil)
        data['Inward (Reverse Charge)'] = clean_number(inward_rc)
        data['IGST'] = clean_number(igst_match)
        data['CGST'] = clean_number(cgst_match)
        data['SGST'] = clean_number(sgst_match)
        
        return data, True
    except Exception as e:
        return str(e), False

def create_consolidated_excel(all_data):
    """Create consolidated Excel file from extracted data"""
    wb = Workbook()
    ws = wb.active
    ws.title = "GSTR-3B Consolidated"
    
    # Styles
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    subheader_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Main header
    ws['A1'] = "GSTR-3B CONSOLIDATED RETURN - FINAL OUTPUT"
    ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    ws.merge_cells('A1:M1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    # Get GSTIN
    gstin = all_data[0].get('GSTIN', 'N/A') if all_data else 'N/A'
    ws['A2'] = f"GSTIN: {gstin}"
    ws['A2'].font = Font(bold=True)
    
    # Column headers
    headers = [
        "Period", "Outward Taxable (Other)", "Outward Taxable (Zero Rated)", 
        "Other Outward (Nil/Exempt)", "Inward (Reverse Charge)", 
        "IGST", "CGST", "SGST"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    ws.row_dimensions[4].height = 30
    
    # Data rows
    for row_num, data in enumerate(all_data, 5):
        row_values = [
            data.get('Period', ''),
            data.get('Outward Taxable (Other)', 0),
            data.get('Outward Taxable (Zero Rated)', 0),
            data.get('Other Outward (Nil/Exempt)', 0),
            data.get('Inward (Reverse Charge)', 0),
            data.get('IGST', 0),
            data.get('CGST', 0),
            data.get('SGST', 0),
        ]
        
        for col_num, value in enumerate(row_values, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            if col_num > 1:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')
    
    # Total row
    total_row = len(all_data) + 5
    ws.cell(row=total_row, column=1).value = "TOTAL"
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    
    for col_num in range(2, len(headers) + 1):
        cell = ws.cell(row=total_row, column=col_num)
        cell.value = f"=SUM({get_column_letter(col_num)}5:{get_column_letter(col_num)}{total_row-1})"
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        cell.border = border
        cell.number_format = '#,##0.00'
    
    # Column widths
    ws.column_dimensions['A'].width = 18
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    return wb

# Main App
st.markdown("""
<div style="text-align: center; padding: 40px 0;">
    <h1 class="header-title">📊 GSTR-3B Consolidator</h1>
    <p class="header-subtitle">Effortlessly consolidate multiple GSTR-3B returns into a single Excel report</p>
</div>
""", unsafe_allow_html=True)

# Sidebar
with st.sidebar:
    st.markdown("### ⚙️ About")
    st.info(
        "**GSTR-3B Consolidator** automatically extracts data from multiple GSTR-3B PDF returns "
        "and creates a unified Excel spreadsheet for easy analysis and reporting."
    )
    
    st.markdown("### 📋 Features")
    features = [
        "✅ Upload multiple GSTR-3B PDFs",
        "✅ Automatic data extraction",
        "✅ Monthly consolidation",
        "✅ Excel export with formatting",
        "✅ Auto-calculated totals",
        "✅ Professional report"
    ]
    for feature in features:
        st.markdown(feature)

# Main content
col1, col2 = st.columns([2, 1])

with col1:
    st.markdown('<div class="upload-section">', unsafe_allow_html=True)
    st.markdown("### 📤 Upload GSTR-3B Returns")
    
    uploaded_files = st.file_uploader(
        "Select one or more GSTR-3B PDF files",
        type=['pdf'],
        accept_multiple_files=True,
        help="Upload PDF files of GSTR-3B returns in any order. The app will automatically consolidate them."
    )
    st.markdown('</div>', unsafe_allow_html=True)

with col2:
    st.markdown('<div class="stats-card">', unsafe_allow_html=True)
    st.markdown('<p class="stats-label">Files Ready</p>', unsafe_allow_html=True)
    st.markdown(f'<p class="stats-value">{len(uploaded_files)}</p>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

if uploaded_files:
    st.markdown("---")
    
    # Processing section
    st.markdown("### 🔄 Processing Files")
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    all_extracted_data = []
    failed_files = []
    
    for idx, uploaded_file in enumerate(uploaded_files):
        status_text.text(f"Processing: {uploaded_file.name}")
        
        data, success = extract_gstr3b_data(uploaded_file)
        
        if success:
            all_extracted_data.append(data)
            status_text.success(f"✅ {uploaded_file.name} - Extracted successfully")
        else:
            failed_files.append((uploaded_file.name, data))
            status_text.error(f"❌ {uploaded_file.name} - Extraction failed")
        
        progress_bar.progress((idx + 1) / len(uploaded_files))
    
    st.markdown("---")
    
    # Results Summary
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown('<div class="stats-card">', unsafe_allow_html=True)
        st.markdown('<p class="stats-label">Files Processed</p>', unsafe_allow_html=True)
        st.markdown(f'<p class="stats-value">{len(all_extracted_data)}</p>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)
    
    with col2:
        st.markdown('<div class="stats-card">', unsafe_allow_html=True)
        st.markdown('<p class="stats-label">Successfully Extracted</p>', unsafe_allow_html=True)
        st.markdown(f'<p class="stats-value">{len(all_extracted_data)}</p>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)
    
    with col3:
        st.markdown('<div class="stats-card">', unsafe_allow_html=True)
        st.markdown('<p class="stats-label">Failed Files</p>', unsafe_allow_html=True)
        st.markdown(f'<p class="stats-value">{len(failed_files)}</p>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)
    
    if failed_files:
        st.markdown('<div class="error-message">', unsafe_allow_html=True)
        st.markdown("### ⚠️ Failed Files")
        for filename, error in failed_files:
            st.markdown(f"**{filename}**: {error}")
        st.markdown('</div>', unsafe_allow_html=True)
    
    if all_extracted_data:
        st.markdown("---")
        
        # Data preview
        st.markdown("### 📊 Data Preview")
        
        preview_df = pd.DataFrame([
            {
                'Period': d.get('Period'),
                'Outward Taxable': f"₹{d.get('Outward Taxable (Other)', 0):,.2f}",
                'IGST': f"₹{d.get('IGST', 0):,.2f}",
                'CGST': f"₹{d.get('CGST', 0):,.2f}",
                'SGST': f"₹{d.get('SGST', 0):,.2f}",
            }
            for d in all_extracted_data
        ])
        
        st.dataframe(preview_df, use_container_width=True, hide_index=True)
        
        st.markdown("---")
        
        # Generate Excel
        st.markdown("### 📥 Generate Report")
        
        if st.button("🎯 Create Consolidated Excel Report", use_container_width=True):
            try:
                wb = create_consolidated_excel(all_extracted_data)
                
                # Save to bytes
                excel_buffer = BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                
                st.markdown('<div class="success-message">', unsafe_allow_html=True)
                st.markdown("✅ **Excel file created successfully!**")
                st.markdown('</div>', unsafe_allow_html=True)
                
                # Download button
                st.download_button(
                    label="⬇️ Download Consolidated Excel",
                    data=excel_buffer.getvalue(),
                    file_name="GSTR3B_Consolidated.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
                # Summary statistics
                st.markdown("### 📈 Summary Statistics")
                
                col1, col2, col3 = st.columns(3)
                
                total_outward = sum(d.get('Outward Taxable (Other)', 0) for d in all_extracted_data)
                total_igst = sum(d.get('IGST', 0) for d in all_extracted_data)
                total_cgst = sum(d.get('CGST', 0) for d in all_extracted_data)
                
                with col1:
                    st.metric("Total Outward Taxable", f"₹{total_outward:,.2f}")
                
                with col2:
                    st.metric("Total IGST", f"₹{total_igst:,.2f}")
                
                with col3:
                    st.metric("Total CGST", f"₹{total_cgst:,.2f}")
                
            except Exception as e:
                st.error(f"❌ Error creating Excel file: {str(e)}")

# Footer
st.markdown("---")
st.markdown("""
<div style="text-align: center; padding: 20px; color: #666; font-size: 0.9em;">
    <p>🔐 <strong>Privacy & Security</strong>: All files are processed locally. No data is stored or sent to external servers.</p>
    <p>📧 <strong>Support</strong>: For issues or feature requests, please contact the development team.</p>
</div>
""", unsafe_allow_html=True)
