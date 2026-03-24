import streamlit as st
import pandas as pd
import io
import os
import re
import json
import sys
import tempfile
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import OrderedDict
from pathlib import Path

# ─── Page Config ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="UK Tool Suite",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── User Store (file-backed so admin changes persist) ────────────────────────
def get_users_file_path():
    if getattr(sys, 'frozen', False):
        return os.path.join(os.path.dirname(sys.executable), "users.json")
    else:
        return os.path.join(os.path.dirname(os.path.abspath(__file__)), "users.json")

USERS_FILE = get_users_file_path()
DEFAULT_USERS = {
    "admin":   {"password": "admin123",    "role": "admin"},
    "article": {"password": "article@2025","role": "user"},
}

def load_users():
    if os.path.exists(USERS_FILE):
        try:
            with open(USERS_FILE, "r") as f:
                return json.load(f)
        except Exception:
            pass
    save_users(DEFAULT_USERS)
    return DEFAULT_USERS

def save_users(users_dict):
    with open(USERS_FILE, "w") as f:
        json.dump(users_dict, f, indent=2)

# ─── Session State Init ───────────────────────────────────────────────────────
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
if "username" not in st.session_state:
    st.session_state.username = ""
if "role" not in st.session_state:
    st.session_state.role = "user"
if "current_page" not in st.session_state:
    st.session_state.current_page = "dashboard"

# ─────────────────────────────────────────────────────────────────────────────
#  CSS: Instagram Stylized
# ─────────────────────────────────────────────────────────────────────────────
DARK_UI_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

:root {
    --side-bg: #000000;
    --main-bg: #ffffff;
    --border: #dbdbdb;
    --primary: #0095f6;
    --text: #262626;
    --text-muted: #8e8e8e;
}

/* Base Styles */
.stApp { background-color: white !important; font-family: 'Inter', sans-serif !important; }
#MainMenu, footer, header { visibility: hidden; }

/* Sidebar Navigation */
section[data-testid="stSidebar"] {
    background-color: var(--side-bg) !important;
    border-right: 1px solid var(--border);
    width: 240px !important;
}

section[data-testid="stSidebar"] [data-testid="stVerticalBlock"] {
    gap: 0px !important;
}

/* Sidebar Branding */
.side-brand {
    padding: 1.5rem 1rem;
    margin-bottom: 0.5rem;
}
.side-brand-text {
    font-size: 1.8rem;
    font-weight: 900;
    color: white;
    letter-spacing: -2px;
}

/* Sidebar Nav Items */
.nav-btn {
    display: flex;
    align-items: center;
    gap: 12px;
    padding: 0.8rem 1rem;
    color: white;
    text-decoration: none;
    border-radius: 8px;
    margin: 4px 0.5rem;
    font-weight: 500;
    transition: all 0.2s;
}
.nav-btn:hover { background: rgba(255,255,255,0.1); }
.nav-btn.active { font-weight: 800; background: rgba(255,255,255,0.1); }

/* Main Content Containers */
.main-container { padding: 1rem 3rem; max-width: 900px; margin: 0 auto; }

/* Utility Cards */
.tool-card {
    background: #ffffff;
    border: 1px solid var(--border);
    border-radius: 12px;
    padding: 1.5rem;
    margin-bottom: 1rem;
    transition: transform 0.2s;
}
.tool-card:hover { transform: translateY(-2px); box-shadow: 0 4px 12px rgba(0,0,0,0.05); }

/* Buttons & Inputs */
.stButton>button {
    background-color: var(--primary) !important;
    color: white !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    width: 100%;
}
.stDownloadButton>button {
    background-color: #262626 !important;
    color: white !important;
    border-radius: 8px !important;
    width: 100%;
}

</style>
"""

# ─────────────────────────────────────────────────────────────────────────────
#  HELPER: Common Logic
# ─────────────────────────────────────────────────────────────────────────────
def get_excel_col_name(n):
    res = ""
    n += 1
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        res = chr(65 + remainder) + res
    return res

# ─────────────────────────────────────────────────────────────────────────────
#  TOOL: GSTR-2A Logic
# ─────────────────────────────────────────────────────────────────────────────
GSTR2A_SHEETS = OrderedDict({
    "B2B": {"header_start": 4, "header_rows": 2},
    "B2BA": {"header_start": 4, "header_rows": 3},
    "CDNR": {"header_start": 4, "header_rows": 2},
    "CDNRA": {"header_start": 4, "header_rows": 3},
    "ECO": {"header_start": 4, "header_rows": 2},
    "ECOA": {"header_start": 4, "header_rows": 3},
    "ISD": {"header_start": 4, "header_rows": 2},
    "ISDA": {"header_start": 4, "header_rows": 3},
    "TDS": {"header_start": 4, "header_rows": 2},
    "TDSA": {"header_start": 4, "header_rows": 2},
    "TCS": {"header_start": 4, "header_rows": 2},
    "IMPG": {"header_start": 4, "header_rows": 2},
    "IMPG SEZ": {"header_start": 4, "header_rows": 2},
})

def format_2a_source(filename):
    basename = os.path.splitext(filename)[0]
    m = re.search(r'_(\d{6})(?:_|$)', basename)
    if m:
        mmyyyy = m.group(1)
        months = {'01':'Jan','02':'Feb','03':'Mar','04':'Apr','05':'May','06':'Jun','07':'Jul','08':'Aug','09':'Sep','10':'Oct','11':'Nov','12':'Dec'}
        month, year = mmyyyy[:2], mmyyyy[2:]
        if month in months: return f"{months[month]} {year}"
    return basename

def build_2a_header(df_raw, config):
    start, num = config["header_start"], config["header_rows"]
    header_rows = [df_raw.iloc[start+i].tolist() if start+i < len(df_raw) else [None]*len(df_raw.columns) for i in range(num)]
    merged = []
    for col_idx in range(len(df_raw.columns)):
        parts = []
        for row in header_rows:
            if col_idx < len(row) and pd.notna(row[col_idx]):
                v = str(row[col_idx]).strip()
                if v and v not in parts: parts.append(v)
        merged.append(" - ".join(parts) if parts else f"Column_{col_idx+1}")
    # Handle duplicates
    seen, unique = {}, []
    for n in merged:
        if n in seen:
            seen[n] += 1
            unique.append(f"{n}_{seen[n]}")
        else:
            seen[n] = 0
            unique.append(n)
    return unique

def extract_2a_data(file_bytes, filename, sheet_name, config):
    try:
        df_raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, header=None, dtype=str)
    except: return None
    headers = build_2a_header(df_raw, config)
    data_start = config["header_start"] + config["header_rows"]
    if data_start >= len(df_raw): return None
    df_data = df_raw.iloc[data_start:].copy()
    df_data.columns = headers[:len(df_data.columns)]
    df_data = df_data.dropna(how='all')
    if df_data.empty: return None
    if sheet_name in {"B2B","B2BA","CDNR","CDNRA","ECO","ECOA"}:
        rate_col = next((c for c in df_data.columns if "Rate" in str(c) and "%" in str(c)), None)
        if rate_col:
            df_data = df_data[df_data[rate_col].astype(str).str.strip() != "-"]
    df_data.insert(0, "Source File", format_2a_source(filename))
    return df_data.reset_index(drop=True)

def consolidate_2a(uploaded_files, selected_sheets, progress_bar=None):
    consolidated = {s: [] for s in selected_sheets}
    total = len(uploaded_files) * len(selected_sheets)
    step = 0
    for f in uploaded_files:
        bytes = f.read(); f.seek(0)
        for s in selected_sheets:
            step += 1
            if progress_bar: progress_bar.progress(step/total)
            df = extract_2a_data(bytes, f.name, s, GSTR2A_SHEETS[s])
            if df is not None: consolidated[s].append(df)
    return {s: pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame() for s, dfs in consolidated.items()}

# ─────────────────────────────────────────────────────────────────────────────
#  TOOL: GSTR-2B Logic
# ─────────────────────────────────────────────────────────────────────────────
GSTR2B_SUMMARY_SHEETS = {"Read me", "ITC Available", "ITC not available", "ITC Reversal", "ITC Rejected"}
B2B_MONTH_MAP = {"01":"January","02":"February","03":"March","04":"April","05":"May","06":"June","07":"July","08":"August","09":"September","10":"October","11":"November","12":"December"}

def get_2b_source(filename):
    stem = Path(filename).stem
    m = re.match(r'^(\d{2})(\d{4})_', stem)
    if m:
        month = B2B_MONTH_MAP.get(m.group(1))
        if month: return f"{month} {m.group(2)}"
    return stem

def find_2b_data_start(ws_rows):
    keys = ["gstin of supplier","gstin of eco","gstin of isd","icegate reference date","invoice number","note number","document number"]
    for i, row in enumerate(ws_rows):
        joined = " ".join([str(v).strip() if v else "" for v in row]).lower()
        if any(k in joined for k in keys): return i + 2
    return 6

def pad_2b_row(row, length):
    lst = list(row)
    if len(lst) < length: lst.extend([None]*(length - len(lst)))
    return lst[:length]

def consolidate_2b(file_paths):
    seen = set(); sheet_names = []
    for fp in file_paths:
        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        for s in wb.sheetnames:
            if s not in GSTR2B_SUMMARY_SHEETS and s not in seen:
                sheet_names.append(s); seen.add(s)
        wb.close()
    
    out_wb = Workbook(); out_wb.remove(out_wb.active)
    stats = {"processed": 0, "sheets": 0, "rows": 0}
    
    for sname in sheet_names:
        max_cols = 0
        for fp in file_paths:
            wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
            if sname in wb.sheetnames:
                if wb[sname].max_column > max_cols: max_cols = wb[sname].max_column
            wb.close()
        if max_cols == 0: max_cols = 1
        
        all_rows = []
        h1, h2 = None, None
        for fp in file_paths:
            src = get_2b_source(os.path.basename(fp))
            wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
            if sname in wb.sheetnames:
                rows = [pad_2b_row(r, max_cols) for r in wb[sname].iter_rows(values_only=True)]
                start = find_2b_data_start(rows)
                if start >= 2 and h1 is None:
                    h1 = rows[start-2]; h2 = rows[start-1]
                for r in rows[start:]:
                    if not all(v is None for v in r): all_rows.append((src, r))
            wb.close()
        
        if not all_rows and h1 is None: continue
        ws = out_wb.create_sheet(title=sname[:31])
        ws.cell(1,1,"Source")
        if h1:
            for i,v in enumerate(h1, 2): ws.cell(1,i,v)
            ws.cell(2,1,None)
            for i,v in enumerate(h2, 2): ws.cell(2,i,v)
            write_ptr = 3
        else: write_ptr = 2
        
        for src, data in all_rows:
            ws.cell(write_ptr, 1, src)
            for i, v in enumerate(data, 2): ws.cell(write_ptr, i, v)
            write_ptr += 1; stats["rows"] += 1
        stats["sheets"] += 1
    stats["processed"] = len(file_paths)
    return out_wb, stats

# ─────────────────────────────────────────────────────────────────────────────
#  TOOL: Excel Reconciliation Logic
# ─────────────────────────────────────────────────────────────────────────────
def clean_recon_vals(series):
    cleaned = series.fillna('').astype(str).str.strip()
    cleaned = cleaned.apply(lambda x: x[:-2] if x.endswith('.0') else x)
    return cleaned.replace({'nan':'','None':'','<NA>':'','NaT':''})

def compare_excels(df1, col1, hr1, df2, col2, hr2):
    s1 = clean_recon_vals(df1[col1].iloc[hr1+1:])
    s2 = clean_recon_vals(df2[col2].iloc[hr2+1:])
    v1 = s1[s1 != '']; v2 = s2[s2 != '']
    c1 = v1.value_counts().to_dict(); c2 = v2.value_counts().to_dict()
    
    st1 = s1.apply(lambda x: 'blank' if x=='' else ('Matched' if c2.get(x,0)==1 else ('Double' if c2.get(x,0)>1 else 'Unmatched')))
    st2 = s2.apply(lambda x: 'blank' if x=='' else ('Matched' if c1.get(x,0)==1 else ('Double' if c1.get(x,0)>1 else 'Unmatched')))
    
    res1, res2 = df1.copy(), df2.copy()
    f1, f2 = pd.Series(['']*len(df1)), pd.Series(['']*len(df2))
    f1.iloc[hr1] = 'Status_Excel1'; f1.iloc[hr1+1:] = st1
    f2.iloc[hr2] = 'Status_Excel2'; f2.iloc[hr2+1:] = st2
    res1.insert(0, 'Status_Excel1', f1); res2.insert(0, 'Status_Excel2', f2)
    return res1, res2

# ─────────────────────────────────────────────────────────────────────────────
#  UI: Login Page
# ─────────────────────────────────────────────────────────────────────────────
def show_login():
    st.markdown("""
    <style>
    .stApp { background-color: #fafafa !important; }
    .login-container { max-width: 350px; margin: 80px auto; padding: 40px; background: white; border: 1px solid #dbdbdb; border-radius: 4px; text-align: center; }
    .login-logo { font-size: 3rem; font-weight: 900; margin-bottom: 20px; letter-spacing: -3px; color: #262626; }
    </style>
    """, unsafe_allow_html=True)
    
    cols = st.columns([1, 1.5, 1])
    with cols[1]:
        st.markdown('<div class="login-container">', unsafe_allow_html=True)
        st.markdown('<div class="login-logo">UK</div>', unsafe_allow_html=True)
        st.markdown('<p style="color:#8e8e8e; font-size: 0.85rem; margin-bottom: 20px;">Professional CA Tool Suite</p>', unsafe_allow_html=True)
        
        user = st.text_input("Username", placeholder="Username", label_visibility="collapsed")
        pw = st.text_input("Password", placeholder="Password", type="password", label_visibility="collapsed")
        
        if st.button("Log in", use_container_width=True):
            users = load_users()
            if user in users and users[user]["password"] == pw:
                st.session_state.logged_in = True
                st.session_state.username = user
                st.session_state.role = users[user]["role"]
                st.rerun()
            else:
                st.error("Invalid credentials.")
        
        st.markdown('<hr style="border-color:#dbdbdb; margin: 20px 0;">', unsafe_allow_html=True)
        st.markdown('<p style="font-size:0.8rem; color:#8e8e8e;">Need access? Contact Admin</p>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
#  UI: Sidebar
# ─────────────────────────────────────────────────────────────────────────────
def show_sidebar():
    with st.sidebar:
        st.markdown('<div class="side-brand"><span class="side-brand-text">UK</span></div>', unsafe_allow_html=True)
        st.markdown(f'<p style="color:#8e8e8e; padding: 0 1rem; font-size: 0.8rem;">Logged in as: <b>{st.session_state.username}</b></p>', unsafe_allow_html=True)
        st.markdown('<hr style="border-color: #262626; margin: 10px 0;">', unsafe_allow_html=True)
        
        pages = {
            "dashboard": ("🏠", "Home"),
            "explore": ("🔍", "Explore"),
            "profile": ("👤", "Profile"),
            "settings": ("⚙️", "Settings") if st.session_state.role == "admin" else None
        }
        
        for key, val in pages.items():
            if val:
                active_class = "active" if st.session_state.current_page == key else ""
                icon, label = val
                if st.button(f"{icon}  {label}", key=f"nav_{key}", use_container_width=True, type="secondary"):
                    st.session_state.current_page = key
                    st.rerun()
        
        st.markdown('<div style="height: 20px;"></div>', unsafe_allow_html=True)
        if st.button("🚪  Logout", key="logout", use_container_width=True):
            st.session_state.logged_in = False
            st.rerun()

# ─────────────────────────────────────────────────────────────────────────────
#  UI: Dashboard
# ─────────────────────────────────────────────────────────────────────────────
def show_dashboard():
    st.markdown('<div class="main-container">', unsafe_allow_html=True)
    st.title(f"Welcome back, {st.session_state.username} 👋")
    st.markdown("---")
    
    st.markdown("### Quick Access")
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("""
        <div class="tool-card">
            <h4>📊 GSTR-2A Consolidator</h4>
            <p style="color:#8e8e8e; font-size:0.85rem;">Consolidate multi-month GSTR-2A returns.</p>
        </div>
        """, unsafe_allow_html=True)
        if st.button("Open GSTR-2A", key="dash_2a"):
            st.session_state.current_page = "gstr2a"; st.rerun()
            
    with col2:
        st.markdown("""
        <div class="tool-card">
            <h4>📈 GSTR-2B Consolidator</h4>
            <p style="color:#8e8e8e; font-size:0.85rem;">Merge GSTR-2B Excel files with source tracking.</p>
        </div>
        """, unsafe_allow_html=True)
        if st.button("Open GSTR-2B", key="dash_2b"):
            st.session_state.current_page = "gstr2b"; st.rerun()

    st.markdown('</div>', unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
#  UI: Explore
# ─────────────────────────────────────────────────────────────────────────────
def show_explore():
    st.markdown('<div class="main-container">', unsafe_allow_html=True)
    st.title("Explore Tools 🔍")
    st.markdown("---")
    
    tools = [
        {"id": "gstr2a", "name": "GSTR-2A Consolidator", "icon": "📊", "desc": "Merge multiple GSTR-2A JSON-derived Excel files into one structured workbook."},
        {"id": "gstr2b", "name": "GSTR-2B Consolidator", "icon": "📈", "desc": "Combine monthly GSTR-2B Excel files with dual-header support and source labels."},
        {"id": "recon", "name": "Excel Reconciliation", "icon": "⚖️", "desc": "Compare two Excel sheets side-by-side to find matches, dublicates and unmatches."},
    ]
    
    for t in tools:
        st.markdown(f"""
        <div class="tool-card">
            <div style="display:flex; align-items:center; gap:15px;">
                <span style="font-size:2rem;">{t['icon']}</span>
                <div>
                    <h4 style="margin:0;">{t['name']}</h4>
                    <p style="margin:0; font-size:0.85rem; color:#8e8e8e;">{t['desc']}</p>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)
        if st.button(f"Launch {t['name']}", key=f"launch_{t['id']}"):
            st.session_state.current_page = t['id']; st.rerun()

    st.markdown('</div>', unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
#  UI: TOOL PAGES
# ─────────────────────────────────────────────────────────────────────────────
def show_gstr2a_page():
    st.markdown('<div class="main-container">', unsafe_allow_html=True)
    st.title("📊 GSTR-2A Consolidator")
    st.markdown("<p style='color:#8e8e8e'>Upload multi-month returns to consolidate</p>", unsafe_allow_html=True)
    
    files = st.file_uploader("Upload GSTR-2A Excel Files", type=['xlsx'], accept_multiple_files=True)
    selected = st.multiselect("Sheets to include", list(GSTR2A_SHEETS.keys()), default=["B2B", "CDNR"])
    
    if st.button("Consolidate GSTR-2A") and files:
        with st.spinner("Processing..."):
            pb = st.progress(0)
            result = consolidate_2a(files, selected, pb)
            
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                for sn, df in result.items():
                    if not df.empty: df.to_excel(writer, sheet_name=sn, index=False)
            output.seek(0)
            
            st.success("Consolidation Complete!")
            st.download_button("Download Merged File", output, "GSTR2A_Merged.xlsx")
    st.markdown('</div>', unsafe_allow_html=True)

def show_gstr2b_page():
    st.markdown('<div class="main-container">', unsafe_allow_html=True)
    st.title("📈 GSTR-2B Consolidator")
    files = st.file_uploader("Upload GSTR-2B Excel Files", type=['xlsx'], accept_multiple_files=True)
    
    if st.button("Consolidate GSTR-2B") and files:
        with tempfile.TemporaryDirectory() as tmpdir:
            paths = []
            for f in files:
                p = os.path.join(tmpdir, f.name)
                with open(p, "wb") as tp: tp.write(f.read())
                paths.append(p)
            
            wb, stats = consolidate_2b(paths)
            buf = io.BytesIO(); wb.save(buf); buf.seek(0)
            
            st.success(f"Merged {stats['processed']} files into {stats['sheets']} sheets.")
            st.download_button("Download GSTR-2B Consolidated", buf, "GSTR2B_Consolidated.xlsx")
    st.markdown('</div>', unsafe_allow_html=True)

def show_recon_page():
    st.markdown('<div class="main-container">', unsafe_allow_html=True)
    st.title("⚖️ Excel Reconciliation Tool")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("### Excel 1")
        f1 = st.file_uploader("Upload Excel 1", type=['xlsx'], key="rf1")
        df1 = None
        if f1:
            xl1 = pd.ExcelFile(f1)
            s1 = st.selectbox("Select Sheet (Excel 1)", xl1.sheet_names, key="rs1")
            df1 = pd.read_excel(xl1, sheet_name=s1, header=None)
            
    with col2:
        st.markdown("### Excel 2")
        f2 = st.file_uploader("Upload Excel 2", type=['xlsx'], key="rf2")
        df2 = None
        if f2:
            xl2 = pd.ExcelFile(f2)
            s2 = st.selectbox("Select Sheet (Excel 2)", xl2.sheet_names, key="rs2")
            df2 = pd.read_excel(xl2, sheet_name=s2, header=None)
    
    if df1 is not None and df2 is not None:
        c1 = [get_excel_col_name(i) for i in range(len(df1.columns))]
        c2 = [get_excel_col_name(i) for i in range(len(df2.columns))]
        
        st.markdown("---")
        col_select1, col_select2 = st.columns(2)
        with col_select1:
            hr1 = st.number_input("Header Row (Excel 1)", 1, len(df1), 1)
            sc1 = st.selectbox("Compare Column (Excel 1)", c1)
        with col_select2:
            hr2 = st.number_input("Header Row (Excel 2)", 1, len(df2), 1)
            sc2 = st.selectbox("Compare Column (Excel 2)", c2)
            
        if st.button("Start Comparison"):
            r1, r2 = compare_excels(df1, c1.index(sc1), hr1-1, df2, c2.index(sc2), hr2-1)
            st.success("Comparison Success!")
            
            out = io.BytesIO()
            with pd.ExcelWriter(out, engine='openpyxl') as writer:
                r1.to_excel(writer, sheet_name="Excel 1", index=False, header=False)
                r2.to_excel(writer, sheet_name="Excel 2", index=False, header=False)
            st.download_button("Download Result", out.getvalue(), "Reconciliation_Result.xlsx")
    st.markdown('</div>', unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
#  UI: Admin / Settings
# ─────────────────────────────────────────────────────────────────────────────
def show_settings():
    st.markdown('<div class="main-container">', unsafe_allow_html=True)
    st.title("Settings & Admin ⚙️")
    
    users = load_users()
    st.markdown("### User Management")
    st.dataframe(pd.DataFrame([{"Username": u, "Role": d["role"]} for u, d in users.items()]), use_container_width=True)
    
    with st.expander("Add New User"):
        nu = st.text_input("New Username")
        np = st.text_input("New Password", type="password")
        nr = st.selectbox("Role", ["user", "admin"])
        if st.button("Add User"):
            users[nu] = {"password": np, "role": nr}
            save_users(users); st.success(f"Added {nu}"); st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
#  ROUTER
# ─────────────────────────────────────────────────────────────────────────────
st.markdown(DARK_UI_CSS, unsafe_allow_html=True)

if not st.session_state.logged_in:
    show_login()
else:
    show_sidebar()
    p = st.session_state.current_page
    if p == "dashboard": show_dashboard()
    elif p == "explore": show_explore()
    elif p == "gstr2a": show_gstr2a_page()
    elif p == "gstr2b": show_gstr2b_page()
    elif p == "recon": show_recon_page()
    elif p == "settings": show_settings()
    elif p == "profile":
        st.markdown('<div class="main-container">', unsafe_allow_html=True)
        st.title("Profile 👤")
        st.info(f"You are logged in as **{st.session_state.username}** with role **{st.session_state.role}**.")
        st.markdown('</div>', unsafe_allow_html=True)
    else: show_dashboard()
